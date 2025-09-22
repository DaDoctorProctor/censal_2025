import os
import re
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------
# Configuration / Globals
# -----------------------
regions = {
    "Frontera": ["027 Nuevo Laredo"],
    "Ribereña": ["007 Camargo", "014 Guerrero", "015 Gustavo Díaz Ordaz", "024 Mier", "025 Miguel Alemán"],
    "Reynosa": ["005 Burgos", "032 Reynosa", "033 Río Bravo", "023 Méndez"],
    "Matamoros": ["010 Cruillas", "035 San Fernando", "022 Matamoros", "040 Valle Hermoso"],
    "Centro": ["001 Abasolo", "008 Casas", "013 Güémez", "016 Hidalgo", "018 Jiménez", "019 Llera", "020 Mainero",
               "030 Padilla", "034 San Carlos", "036 San Nicolás", "037 Soto la Marina", "041 Victoria", "042 Villagrán"],
    "Mante": ["004 Antiguo Morelos", "006 Bustamante", "017 Jaumave", "026 Miquihuana", "031 Palmillas", "039 Tula",
              "021 El Mante", "011 Gómez Farías", "012 González", "028 Nuevo Morelos", "029 Ocampo", "043 Xicoténcatl"],
    "Sur": ["002 Aldama", "003 Altamira", "009 Ciudad Madero", "038 Tampico"]
}

YEARS = [2008, 2013, 2018, 2023]

INPUT_CSV = "input/SAIC_Exporta_Clean.csv"
OUTPUT_DIR = "output"

# ---------- Helper functions ----------
def sanitize_filename(code_and_name):
    # produce safe, verbose filename: replace spaces/punctuation with underscores, keep ascii
    s = re.sub(r'[^\w\s-]', '', code_and_name)  # remove punctuation
    s = re.sub(r'\s+', '_', s.strip())
    return s

def clean_sectores_column(df):
    # Remove any "Sector XX" or "Sector XX-YY" pattern in 'Actividad económica'
    df["Actividad económica"] = df["Actividad económica"].apply(
        lambda x: re.sub(r'Sector\s*\d+(-\d+)?\s*', '', str(x))
    )
    return df

def add_total_row(df):
    # Sum numeric-only columns, keep non-numeric columns as-is. Place "Total" in Sectores column.
    numeric_sum = df.select_dtypes(include='number').sum()
    # create a row full of NaNs, then fill numeric sums
    total_row = {col: (numeric_sum[col] if col in numeric_sum.index else "") for col in df.columns}
    total_row["Sectores"] = "Total"
    # convert to DataFrame and append
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

def process_variable(df, variable, regions):
    """
    Build region-year aggregated pivot for the variable
    """
    results = []

    for region, municipios in regions.items():
        temp = (
            df[df["Municipio"].isin(municipios) & (df["Municipio"] != "")]
            .groupby(["Año Censal", "Actividad económica"], as_index=False)[variable]
            .sum()
        )
        temp["Region"] = region
        results.append(temp)

    combined = pd.concat(results, ignore_index=True) if results else pd.DataFrame(columns=["Año Censal", "Actividad económica", variable, "Region"])

    pivoted = combined.pivot_table(
        index="Actividad económica",
        columns=["Region", "Año Censal"],
        values=variable,
        aggfunc="sum",
        fill_value=0
    ) if not combined.empty else pd.DataFrame()

    # Ensure all columns exist in order
    col_tuples = []
    for region in regions.keys():
        for year in YEARS:
            col_tuples.append((region, year))

    if pivoted.empty:
        activities = df["Actividad económica"].unique().tolist()
        pivoted = pd.DataFrame(0, index=activities, columns=pd.MultiIndex.from_tuples(col_tuples))
    else:
        for tup in col_tuples:
            if tup not in pivoted.columns:
                pivoted[tup] = 0

    pivoted = pivoted[[ (region, year) for region in regions.keys() for year in YEARS ]]
    pivoted.columns = [f"{region}_{year}" for region, year in pivoted.columns]
    pivoted.reset_index(inplace=True)

    # ✅ Keep original label for first column
    pivoted.rename(columns={"Actividad económica": "Actividad económica"}, inplace=True)

    return pivoted

# ---------- Excel formatting utilities (based on your original logic) ----------
def format_numbers_with_commas_all_sheets(filename):
    wb = load_workbook(filename)

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                # Apply number format to all cells in numeric columns, including formulas
                cell.number_format = '#,##0'

    wb.save(filename)

def add_two_row_header(filename, sheet_name, regions_local):
    wb = load_workbook(filename)
    ws = wb[sheet_name]

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Shift all data down by 2 rows
    ws.insert_rows(1, amount=2)

    # Merge "Sectores" A1:A2 and center + bold
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    cell = ws.cell(row=1, column=1, value="Sectores")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True)
    cell.border = thin_border

    # Add region names in row 1, merge over len(YEARS) columns, bold + centered
    col = 2
    for region in regions_local.keys():
        start_col = col
        end_col = col + len(YEARS) - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        region_cell = ws.cell(row=1, column=start_col, value=region)
        region_cell.alignment = Alignment(horizontal="center", vertical="center")
        region_cell.font = Font(bold=True)
        region_cell.border = thin_border

        # Write years in row 2, bold + centered
        for i, year in enumerate(YEARS):
            year_cell = ws.cell(row=2, column=start_col + i, value=year)
            year_cell.alignment = Alignment(horizontal="center", vertical="center")
            year_cell.font = Font(bold=True)
            year_cell.border = thin_border
        col += len(YEARS)

    # Delete the (old) third row which contained the header from to_excel
    try:
        ws.delete_rows(3)
    except Exception:
        pass

    # Center all cells except first column and apply borders
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=2, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Add borders + bold handling for first column
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.border = thin_border

    # Bold the last row ("Total") if present
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)

    wb.save(filename)

def wrap_text_first_column(filename, sheet_name):
    wb = load_workbook(filename)
    ws = wb[sheet_name]

    # Apply wrap text to the first column
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(filename)

def adjust_first_column_and_autofit_rows(filename, sheet_name, first_col_width=50, line_height=15, padding=2):
    wb = load_workbook(filename)
    ws = wb[sheet_name]

    # 1️⃣ Adjust first column width
    ws.column_dimensions[get_column_letter(1)].width = first_col_width

    # 2️⃣ Estimate column widths for all columns
    col_widths = {}
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        col_letter = get_column_letter(col[0].column)
        if col[0].column == 1:
            col_widths[col[0].column] = first_col_width
        else:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_widths[col[0].column] = max_len + padding

    # 3️⃣ Adjust row heights
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        max_lines = 1
        for cell in row:
            if cell.value:
                col_w = col_widths.get(cell.column, 10)
                lines = math.ceil(len(str(cell.value)) / max(col_w, 1))
                if lines > max_lines:
                    max_lines = lines
        ws.row_dimensions[row[0].row].height = max_lines * line_height

    wb.save(filename)

def split_by_region_with_links(filename, base_sheet, regions_local, years):
    wb = load_workbook(filename)
    ws = wb[base_sheet]

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    header_row_region = 1
    header_row_years = 2
    data_start_row = 3
    data_end_row = ws.max_row

    # Propagate region names for merged cells
    col_region_map = {}
    max_col = ws.max_column
    current_region = None
    for col in range(2, max_col + 1):
        val = ws.cell(row=header_row_region, column=col).value
        if val is not None:
            current_region = val
        col_region_map[col] = current_region

    # Map columns per region per year
    region_year_col = {}
    for col in range(2, max_col + 1):
        region_name = col_region_map[col]
        year_val = ws.cell(row=header_row_years, column=col).value
        if region_name not in region_year_col:
            region_year_col[region_name] = {}
        region_year_col[region_name][year_val] = col

    # Create tabs per region
    for region in regions_local.keys():
        new_sheet_name = f"{base_sheet}_{region}"
        if new_sheet_name in wb.sheetnames:
            del wb[new_sheet_name]
        ws_new = wb.create_sheet(new_sheet_name)

        # --- Two-row header ---
        ws_new.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        c = ws_new.cell(row=1, column=1, value="Sectores")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

        start_col = 2
        end_col = start_col + len(years) - 1
        ws_new.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        c = ws_new.cell(row=1, column=start_col, value=region)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

        for i, year in enumerate(years):
            yc = ws_new.cell(row=2, column=start_col+i, value=year)
            yc.font = Font(bold=True)
            yc.alignment = Alignment(horizontal="center", vertical="center")
            yc.border = thin_border

        # --- Copy data (with formulas linking to base sheet) ---
        for r in range(data_start_row, data_end_row + 1):
            # Sectores
            src_val = ws.cell(row=r, column=1).value
            dest_cell = ws_new.cell(row=r, column=1, value=src_val)
            dest_cell.alignment = Alignment(wrap_text=True, vertical="top")
            dest_cell.border = thin_border

            # Region-year formulas
            for i, year in enumerate(years):
                col = region_year_col.get(region, {}).get(year)
                if col is not None:
                    formula = f"='{base_sheet}'!{get_column_letter(col)}{r}"
                    dest_cell = ws_new.cell(row=r, column=start_col+i, value=formula)
                    dest_cell.alignment = Alignment(horizontal="center", vertical="center")
                    dest_cell.border = thin_border
                else:
                    # place zero if mapping not found
                    dest_cell = ws_new.cell(row=r, column=start_col+i, value=0)
                    dest_cell.alignment = Alignment(horizontal="center", vertical="center")
                    dest_cell.border = thin_border

        # Bold last row (Total)
        last_row = ws_new.max_row
        for cell in ws_new[last_row]:
            cell.font = Font(bold=True)

        # Adjust Sectores column width
        ws_new.column_dimensions[get_column_letter(1)].width = 50

    wb.save(filename)

# ---------- Nation/State wide CSV export ----------
def export_nation_state_wide_csvs(df, years, output_dir):
    csv_dir = os.path.join(output_dir, "csv")
    os.makedirs(csv_dir, exist_ok=True)

    for entidad, label in [("00 Total Nacional", "00_Total_Nacional"), ("28 Tamaulipas", "28_Tamaulipas")]:
        df_filtered = df[(df["Entidad"] == entidad) & (df["Municipio"].str.strip() == "")]
        if df_filtered.empty:
            continue

        # Keep only years >= 2008
        df_filtered = df_filtered[df_filtered["Año Censal"].isin(years)]

        # Determine variable columns
        required = ["Año Censal", "Entidad", "Municipio", "Actividad económica"]
        var_cols = [c for c in df_filtered.columns if c not in required]

        # Build wide-format DataFrame
        pivot_list = []
        for var in var_cols:
            temp = df_filtered.pivot_table(
                index="Actividad económica",
                columns="Año Censal",
                values=var,
                aggfunc="sum",
                fill_value=0
            )
            temp.columns = [f"{var}_{int(col)}" for col in temp.columns]
            pivot_list.append(temp)

        df_wide = pd.concat(pivot_list, axis=1)
        df_wide.reset_index(inplace=True)  # keep 'Actividad económica'

        # Clean "Sector XX" from names
        df_wide["Actividad económica"] = df_wide["Actividad económica"].apply(
            lambda x: re.sub(r'Sector\s*\d+(-\d+)?\s*', '', str(x))
        )

        # Save CSV
        filename_safe = sanitize_filename(label)
        csv_path = os.path.join(csv_dir, f"{filename_safe}.csv")
        df_wide.to_csv(csv_path, index=False)
        print(f"Saved wide-format CSV: {csv_path}")


# ---------------- Main pipeline ----------------
def main():
    # Create output folders
    csv_dir = os.path.join(OUTPUT_DIR, "csv")
    excel_dir = os.path.join(OUTPUT_DIR, "excel")
    os.makedirs(csv_dir, exist_ok=True)
    os.makedirs(excel_dir, exist_ok=True)

    # Read CSV
    df = pd.read_csv(INPUT_CSV, dtype=str)
    df.columns = [c.strip() for c in df.columns]

    # Convert Municipio to string
    df["Municipio"] = df["Municipio"].fillna("").astype(str).str.strip()

    # Keep only years >= 2008
    df["Año Censal"] = pd.to_numeric(df["Año Censal"], errors='coerce')
    df = df[df["Año Censal"] >= 2008]

    # Adjust YEARS dynamically
    global YEARS
    YEARS = [year for year in YEARS if year >= 2008]

    # ---------------- Nation/State wide CSVs ----------------
    export_nation_state_wide_csvs(df, YEARS, OUTPUT_DIR)

    # Ensure required identifier columns exist
    required = ["Año Censal", "Entidad", "Municipio", "Actividad económica"]
    for col in required:
        if col not in df.columns:
            raise ValueError(f"Expected column '{col}' in input CSV")

    # Determine variable columns
    all_cols = list(df.columns)
    var_cols = [c for c in all_cols if c not in required]

    # Convert variable columns to numeric
    for v in var_cols:
        df[v] = pd.to_numeric(df[v].str.replace(',', '').str.strip(), errors='coerce').fillna(0)

    # ---------------- Regional variable CSVs & Excel ----------------
    for var in var_cols:
        print(f"Processing variable: {var}")

        pivot = process_variable(df, var, regions)
        pivot = clean_sectores_column(pivot)
        pivot = add_total_row(pivot)

        m = re.match(r'^(\w+)\s*(.*)$', var)
        if m:
            code = m.group(1)
            name = m.group(2)
        else:
            code = var
            name = ""
        verbose = f"{code}_{name}".strip()
        filename_safe = sanitize_filename(verbose)

        # ---------------- CSV Output ----------------
        csv_path = os.path.join(csv_dir, f"{filename_safe}.csv")
        pivot.to_csv(csv_path, index=False)
        print(f"Saved CSV: {csv_path}")

        # ---------------- Excel Output ----------------
        excel_path = os.path.join(excel_dir, f"{filename_safe}.xlsx")
        base_sheet_name = code
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            pivot.to_excel(writer, sheet_name=base_sheet_name, index=False)

        add_two_row_header(excel_path, base_sheet_name, regions)
        wrap_text_first_column(excel_path, base_sheet_name)
        split_by_region_with_links(excel_path, base_sheet_name, regions, YEARS)
        adjust_first_column_and_autofit_rows(excel_path, base_sheet_name, first_col_width=50, line_height=15)

        for region in regions.keys():
            sheet_region = f"{base_sheet_name}_{region}"
            adjust_first_column_and_autofit_rows(excel_path, sheet_region, first_col_width=50, line_height=15)
            wrap_text_first_column(excel_path, sheet_region)

        format_numbers_with_commas_all_sheets(excel_path)
        print(f"Saved Excel: {excel_path}")

    print("✅ All variables processed.")
    print("✅ Nation and Tamaulipas CSVs exported to 'csv/' in wide pivot format")
    print("✅ Regional variable CSVs exported to 'csv/' and Excels to 'excel/'")


if __name__ == "__main__":
    main()
